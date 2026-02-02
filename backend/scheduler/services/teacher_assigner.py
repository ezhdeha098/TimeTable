"""
Teacher Assignment Service
Assigns teachers to timetable slots based on uploaded preferences.
"""
from typing import List, Dict, Set, Tuple
from django.db import transaction
from ..models import (
    Teacher,
    TeacherPreference,
    TimetableSlot,
    Subject,
    TimeSlot,
)


class TeacherAssignmentError(Exception):
    """Raised when teacher assignment fails"""
    pass


class TeacherAssigner:
    """
    Assigns teachers to already-generated timetable slots.
    
    Assignment Priority:
    1. Specific course + specific type (e.g., CS201 + Theory)
    2. Specific course + wildcard type (e.g., CS201 + *)
    3. Wildcard course + specific type (e.g., * + Theory)
    4. Wildcard course + wildcard type (e.g., * + *)
    """
    
    def __init__(self):
        self.teachers: Dict[int, Teacher] = {}
        self.preferences: List[TeacherPreference] = []
        self.timetable_slots: List[TimetableSlot] = []
        self.subject_map: Dict[int, Subject] = {}
        
    def load_data(self):
        """Load all necessary data from database"""
        # Load all teachers and their preferences
        self.teachers = {t.id: t for t in Teacher.objects.all()}
        self.preferences = list(TeacherPreference.objects.select_related('teacher').all())
        
        # Load timetable slots that need assignment
        self.timetable_slots = list(
            TimetableSlot.objects.select_related('subject', 'section', 'timeslot')
            .filter(teacher__isnull=True)
            .all()
        )
        
        # Load subjects for quick lookup
        self.subject_map = {s.id: s for s in Subject.objects.all()}
        
    def assign_teachers(self, clear_existing: bool = False) -> Dict:
        """
        Main assignment logic.
        
        Args:
            clear_existing: If True, clear all existing teacher assignments first
            
        Returns:
            Dict with assignment results and statistics
        """
        if clear_existing:
            TimetableSlot.objects.all().update(teacher=None)
            
        self.load_data()
        
        if not self.timetable_slots:
            return {
                'status': 'no-slots',
                'message': 'No unassigned timetable slots found',
                'assigned': 0,
                'unassigned': 0,
            }
            
        if not self.preferences:
            return {
                'status': 'no-preferences',
                'message': 'No teacher preferences uploaded',
                'assigned': 0,
                'unassigned': len(self.timetable_slots),
            }
        
        # Sort preferences by specificity (most specific first)
        sorted_prefs = self._sort_preferences_by_priority(self.preferences)
        
        # Track assignments
        assigned_slots: Set[int] = set()
        teacher_workload: Dict[int, int] = {t_id: 0 for t_id in self.teachers.keys()}
        teacher_timeslots: Dict[int, Set[int]] = {t_id: set() for t_id in self.teachers.keys()}
        
        assignments_to_save = []
        
        # Process each preference in order
        for pref in sorted_prefs:
            teacher_id = pref.teacher.id
            sections_needed = pref.sections_count
            
            # Find matching unassigned slots
            matching_slots = self._find_matching_slots(
                pref,
                assigned_slots,
                teacher_timeslots[teacher_id]
            )
            
            # Assign up to sections_needed
            assigned_count = 0
            for slot in matching_slots:
                if assigned_count >= sections_needed:
                    break
                    
                # Check if teacher has conflict at this timeslot
                if slot.timeslot.id in teacher_timeslots[teacher_id]:
                    continue
                    
                # Assign teacher to slot
                slot.teacher_id = teacher_id
                assignments_to_save.append(slot)
                assigned_slots.add(slot.id)
                teacher_timeslots[teacher_id].add(slot.timeslot.id)
                teacher_workload[teacher_id] += 1
                assigned_count += 1
        
        # Save all assignments in bulk
        with transaction.atomic():
            for slot in assignments_to_save:
                slot.save(update_fields=['teacher'])
        
        # Calculate statistics
        total_slots = len(self.timetable_slots)
        assigned = len(assigned_slots)
        unassigned = total_slots - assigned
        
        # Generate warnings
        warnings = []
        for t_id, workload in teacher_workload.items():
            if workload > 0:
                teacher = self.teachers[t_id]
                warnings.append(f"{teacher.name}: {workload} slots assigned")
        
        # Check for unassigned slots
        if unassigned > 0:
            warnings.append(f"{unassigned} slots remain unassigned (no matching teachers)")
        
        return {
            'status': 'ok',
            'assigned': assigned,
            'unassigned': unassigned,
            'total_slots': total_slots,
            'teacher_workloads': teacher_workload,
            'warnings': warnings,
        }
    
    def _sort_preferences_by_priority(self, prefs: List[TeacherPreference]) -> List[TeacherPreference]:
        """
        Sort preferences by specificity (most specific first).
        
        Priority levels:
        1. Specific course + specific type (both can_theory and can_lab are not both True)
        2. Specific course + any type (both can_theory and can_lab are True)
        3. Any course + specific type
        4. Any course + any type
        """
        def priority_key(pref: TeacherPreference) -> Tuple[int, str, str]:
            is_wildcard_course = pref.course_code == '*'
            is_wildcard_type = pref.can_theory and pref.can_lab
            
            if not is_wildcard_course and not is_wildcard_type:
                priority = 0  # Most specific
            elif not is_wildcard_course and is_wildcard_type:
                priority = 1
            elif is_wildcard_course and not is_wildcard_type:
                priority = 2
            else:
                priority = 3  # Least specific
                
            # Secondary sort: by course code, then teacher name
            return (priority, pref.course_code, pref.teacher.name)
        
        return sorted(prefs, key=priority_key)
    
    def _find_matching_slots(
        self,
        pref: TeacherPreference,
        assigned_slots: Set[int],
        teacher_existing_timeslots: Set[int]
    ) -> List[TimetableSlot]:
        """
        Find timetable slots that match the teacher's preference.
        
        Args:
            pref: Teacher preference to match
            assigned_slots: Set of slot IDs already assigned
            teacher_existing_timeslots: Set of timeslot IDs already used by this teacher
            
        Returns:
            List of matching unassigned TimetableSlot objects
        """
        matching = []
        
        for slot in self.timetable_slots:
            # Skip already assigned slots
            if slot.id in assigned_slots:
                continue
            
            subject = self.subject_map[slot.subject_id]
            
            # Check course match
            if pref.course_code != '*' and subject.code != pref.course_code:
                continue
            
            # Check type match (theory vs lab)
            if subject.is_lab:
                if not pref.can_lab:
                    continue
            else:
                if not pref.can_theory:
                    continue
            
            matching.append(slot)
        
        # Group by section to respect "sections_count" properly
        # Each section typically has 2 theory slots or 1 lab slot per week
        # We want to assign complete sections, not individual slots
        
        # Group slots by (section_id, subject_id)
        from collections import defaultdict
        section_groups = defaultdict(list)
        for slot in matching:
            key = (slot.section_id, slot.subject_id)
            section_groups[key].append(slot)
        
        # Flatten back to list, but prioritize complete section assignments
        result = []
        for section_slots in section_groups.values():
            result.extend(section_slots)
        
        return result


def import_teachers_from_excel(excel_file, clear_existing: bool = True) -> Dict:
    """
    Import teacher preferences from Excel file.
    
    Expected columns:
    - Teacher Name
    - Course Code (can be '*' for any course)
    - Sections Count (integer)
    - Type (Theory/Lab/* for both)
    
    Args:
        excel_file: Uploaded file object
        clear_existing: If True, delete existing teachers and preferences first
        
    Returns:
        Dict with import results
    """
    import openpyxl
    from io import BytesIO
    
    # Read Excel file
    wb = openpyxl.load_workbook(BytesIO(excel_file.read()))
    ws = wb.active
    
    # Parse header
    headers = [cell.value for cell in ws[1]]
    expected = ['Teacher Name', 'Course Code', 'Sections Count', 'Type']
    
    # Flexible header matching (case-insensitive, strip spaces)
    header_map = {}
    for i, h in enumerate(headers):
        if h:
            h_clean = str(h).strip().lower()
            for exp in expected:
                if exp.lower() in h_clean or h_clean in exp.lower():
                    header_map[exp] = i
                    break
    
    missing = [h for h in expected if h not in header_map]
    if missing:
        raise TeacherAssignmentError(f"Missing required columns: {', '.join(missing)}")
    
    # Parse rows
    teacher_data = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):  # Skip empty rows
            continue
            
        try:
            teacher_name = str(row[header_map['Teacher Name']]).strip()
            course_code = str(row[header_map['Course Code']]).strip()
            sections_count = int(row[header_map['Sections Count']])
            type_str = str(row[header_map['Type']]).strip().lower()
            
            # Parse type
            if type_str in ['theory', 't']:
                can_theory, can_lab = True, False
            elif type_str in ['lab', 'l']:
                can_theory, can_lab = False, True
            elif type_str in ['*', 'both', 'any']:
                can_theory, can_lab = True, True
            else:
                raise ValueError(f"Invalid Type value: {type_str}")
            
            teacher_data.append({
                'name': teacher_name,
                'course_code': course_code,
                'sections_count': sections_count,
                'can_theory': can_theory,
                'can_lab': can_lab,
            })
        except Exception as e:
            raise TeacherAssignmentError(f"Error parsing row {row_idx}: {str(e)}")
    
    if not teacher_data:
        raise TeacherAssignmentError("No valid teacher data found in Excel file")
    
    # Save to database
    with transaction.atomic():
        if clear_existing:
            TeacherPreference.objects.all().delete()
            Teacher.objects.all().delete()
        
        # Create teachers and preferences
        teachers_created = 0
        preferences_created = 0
        
        for data in teacher_data:
            # Get or create teacher
            teacher, created = Teacher.objects.get_or_create(name=data['name'])
            if created:
                teachers_created += 1
            
            # Create preference
            TeacherPreference.objects.create(
                teacher=teacher,
                course_code=data['course_code'],
                sections_count=data['sections_count'],
                can_theory=data['can_theory'],
                can_lab=data['can_lab'],
            )
            preferences_created += 1
    
    return {
        'status': 'ok',
        'teachers_created': teachers_created,
        'preferences_created': preferences_created,
        'total_rows': len(teacher_data),
    }
